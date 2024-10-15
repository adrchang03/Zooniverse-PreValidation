import pandas as pd
from tkinter import Tk, filedialog, Toplevel, Label, StringVar
import openpyxl
import time
from threading import Thread

# Step 1: Request the user to select an Excel file through file explorer
def select_file():
    root = Tk()
    root.withdraw()  # Hide the main Tkinter window
    file_path = filedialog.askopenfilename(title="Select an Excel file", 
                                           filetypes=[("Excel files", "*.xlsx *.xls")])
    return file_path

# Step 2: Create the specified headings
def create_headings(file_path):
    # Load workbook and active sheet
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    # Define the column headers
    headers = ["VALIDATION", "SEASON", "FILENAME", "SPECIES 1", "HOW MANY OF SPECIES 1",
               "SPECIES 2", "HOW MANY OF SPECIES 2", "TIME OF DAY", "TEMPERATURE",
               "MONTH", "HABITAT"]

    # Insert the headers in the first row (A1 to K1)
    for col, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col).value = header

    # Save the workbook
    wb.save(file_path)

# Helper function to find the first row where columns A to K are empty
def find_next_empty_row(sheet):
    for row in range(2, sheet.max_row + 2):  # Start at row 2
        if all(sheet.cell(row=row, column=col).value is None for col in range(1, 12)):  # A to K
            return row
    return sheet.max_row + 1  # If no empty row found, return the next new row

# Extract season from the filename
def extract_season(filename):
    try:
        # Split by underscores, grab the second part, then remove the file extension
        season_part = filename.split('_')[1]
        return season_part.split(' (')[0]  # Remove extension if present
    except IndexError:
        return "Unknown"

# Get habitat type based on the filename
def get_habitat_type(filename):
    try:
        if "CPW" in filename:
            # Extract the two characters following 'CPW'
            habitat_code = filename[filename.index("CPW") + 3:filename.index("CPW") + 5]
            # Map habitat codes to habitat types
            if habitat_code in ["01", "05"]:
                return "Coastal Sage Scrub"
            elif habitat_code in ["02", "03", "06", "09", "10", "11", "12", "15", "17"]:
                return "Disturbed"
            elif habitat_code in ["04", "08", "16", "18"]:
                return "Riparian"
            elif habitat_code in ["07", "13"]:
                return "Disturbed Coastal Sage Scrub"
            elif habitat_code in ["19", "20"]:
                return "Oak Woodland"
        return "MISSING HABITAT DATA"
    except Exception:
        return "MISSING HABITAT DATA"


# Step 3: Process the data in column L and check validation
def process_data(file_path, progress_var, update_progress_callback):
    # Load the Excel file into a pandas DataFrame
    df = pd.read_excel(file_path)

    # Step 3a: Read through column L (filenames) and group the data by filenames
    grouped = df.groupby(df.columns[11])  # Column L (12th column is index 11)

    # Load workbook with openpyxl to write results
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    # Track unique filenames that have already been written
    written_filenames = set()

    # Find the first empty row for output based on columns A to K
    next_row = find_next_empty_row(sheet)

    total_files = len(grouped)  # Total number of files to process
    current_file = 0  # To track progress

    for filename, group in grouped:
        current_file += 1
        # Update progress via callback to ensure it's done in the main thread
        update_progress_callback(current_file, total_files)

        # Ensure unique filenames are written only once
        if filename not in written_filenames:
            written_filenames.add(filename)

            # Step 3b: Compare columns M, N, O, P (13th to 16th columns: index 12, 13, 14, 15)
            is_valid = group.iloc[:, [12, 13, 14, 15]].nunique().eq(1).all()

            # Step 3c: Write the filename to column C (only once per filename)
            sheet.cell(row=next_row, column=3).value = filename  # Column C (3rd column)

            # Step 3d: Write TRUE if valid, FALSE if invalid in column A (1st column)
            sheet.cell(row=next_row, column=1).value = "TRUE" if is_valid else "FALSE"

            # Step 4: Extract season from filename and write in column B
            season = extract_season(filename)
            sheet.cell(row=next_row, column=2).value = season  # Column B (2nd column)

            # Step 5: If valid, copy columns M-T (source) to D-K (destination)
            if is_valid:
                # Correct range is columns M-T (13 to 20 in DataFrame) to D-K (4 to 11 in sheet)
                for i, col_index in enumerate(range(13, 21), start=4):  # Copy from columns M-T to D-K
                    sheet.cell(row=next_row, column=i).value = group.iloc[0, col_index - 1]  # Copy from row

        # Step 6: If invalid (FALSE), perform additional checks on column Q (index 16)
            else:
                # Filter rows where column Q (index 16) does not have "NODATA"
                non_nodata_rows = group[group.iloc[:, 16] != "NODATA"]

                if non_nodata_rows.empty:
                    # If all rows for this filename have "NODATA", set H, I, J to "NODATA"
                    sheet.cell(row=next_row, column=8).value = "NODATA"  # Column I (9th column)
                    sheet.cell(row=9, column=9).value = "NODATA"         # Column J (10th column)
                    sheet.cell(row=next_row, column=10).value = "NODATA"  # Column K (11th column)
                else:
                    # Use the first row that doesn't have "NODATA"
                    valid_row = non_nodata_rows.iloc[0]

                    # Copy values from columns Q, R, S (16th to 18th index) to columns I, J, K (9th to 11th column)
                    sheet.cell(row=next_row, column=8).value = valid_row.iloc[16]  # Column Q to I
                    sheet.cell(row=next_row, column=9).value = valid_row.iloc[17]  # Column R to J
                    sheet.cell(row=next_row, column=10).value = valid_row.iloc[18]  # Column S to K


            # Move to the next row for the next filename
            next_row = find_next_empty_row(sheet)

    # Save the modified workbook
    wb.save(file_path)

# Step 6: Create a loading window with a progress message
def show_loading_window(progress_var):
    loading_window = Toplevel()
    loading_window.title("Processing...")
    label = Label(loading_window, textvariable=progress_var, padx=20, pady=20)
    label.pack()
    return loading_window

# Step 7: Add habitat type to columns K and T based on filename
def add_habitat_type(file_path):
    # Load workbook and active sheet
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    # Find the max row in both Column C and Column L
    max_row_c = sheet.max_row  # Assume column C is longest initially
    max_row_l = max(sheet.max_row, len([row for row in sheet.iter_rows(min_col=12, max_col=12, values_only=True) if row[0]]))  # Check length of column L

    max_row = max(max_row_c, max_row_l)

    # Process Column C for habitat type (output to column K)
    for row in range(2, max_row + 1):  # Start from row 2 to avoid the header
        filename_col_c = sheet.cell(row=row, column=3).value  # Column C (3rd column)
        if filename_col_c:
            habitat_type_col_k = get_habitat_type(filename_col_c)
            sheet.cell(row=row, column=11).value = habitat_type_col_k  # Column K (11th column)

    # Process Column L for habitat type (output to column T)
    for row in range(2, max_row + 1):  # Start from row 2 to avoid the header
        filename_col_l = sheet.cell(row=row, column=12).value  # Column L (12th column)
        if filename_col_l:
            habitat_type_col_t = get_habitat_type(filename_col_l)
            sheet.cell(row=row, column=20).value = habitat_type_col_t  # Column T (20th column)

    # Save the workbook with the newly added habitat types
    wb.save(file_path)




# Function to process the Excel file in a separate thread
def process_excel_in_thread(file_path, progress_var, root):
    def update_progress(current_file, total_files):
        root.after(0, progress_var.set, f"Processing file {current_file}/{total_files}...")

    # Create headings first
    create_headings(file_path)

    # Process the file and update the progress
    process_data(file_path, progress_var, update_progress)

    # Call the function to add habitat type
    add_habitat_type(file_path)


    # Close the loading window after completion
    root.quit()  # Close the Tkinter loop

if __name__ == "__main__":
    file_path = select_file()

    if file_path:
        # Initialize the main Tkinter window
        root = Tk()
        root.withdraw()  # Hide the main window

        # Progress tracking variable
        progress_var = StringVar()
        progress_var.set("Starting...")

        # Create and show the loading window
        loading_window = show_loading_window(progress_var)

        # Run the Excel processing in a separate thread
        processing_thread = Thread(target=process_excel_in_thread, args=(file_path, progress_var, root))
        processing_thread.start()

        # Start the Tkinter event loop to show the loading window
        root.mainloop()

        print("Excel file modified successfully!")
    else:
        print("No file selected.")
